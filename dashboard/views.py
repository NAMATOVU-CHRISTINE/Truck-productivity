import csv
from django.http import HttpResponse, Http404
from django.contrib.auth.decorators import login_required
@login_required
def download_report(request, upload_id):
    """Download processed TruckPerformanceData as CSV for a given upload."""
    upload = get_object_or_404(CSVUpload, id=upload_id, processed=True)
    data_qs = TruckPerformanceData.objects.filter(csv_upload=upload)
    # Filter out unwanted rows for the report as well
    data_qs = data_qs.exclude(driver_name='Unknown Driver', truck_number='TRUCK_999') \
                   .exclude(customer_name='Unknown Customer') \
                   .exclude(current_status='Pending Departure')
    if not data_qs.exists():
        raise Http404("No processed data found for this upload.")
    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f"report_{upload.name.replace(' ', '_')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    # Write header
    fields = [
        'create_date', 'month_name', 'transporter', 'load_number', 'driver_name', 'truck_number', 'customer_name',
        'dj_departure_time', 'clockin_time', 'arrival_at_customer', 'arrival_at_depot', 'days_spent',
        'departure_time_from_customer', 'total_distance', 'total_time', 'delivery_time', 'efficiency_score', 'current_status'
    ]
    # Write header with user-friendly names
    header = [
        'Create Date', 'Month', 'Transporter', 'Load Number', 'Driver Name', 'Truck Number', 'Customer Name',
        'DJ Departure Time', 'Clock-in Time', 'Arrival at Customer', 'Arrival at Depot', 'Days Spent',
        'Departure Time from Customer', 'Total Distance', 'Total Time', 'Delivery Time', 'Efficiency Score', 'Current Status'
    ]
    writer.writerow(header)
    # Write data rows
    for obj in data_qs:
        # Calculate days spent if possible
        arrival_at_customer = getattr(obj, 'arrival_at_customer', None)
        arrival_at_depot = getattr(obj, 'arrival_at_depot', None)
        days_spent = ''
        if obj.dj_departure_time and obj.arrival_at_depot:
            days_spent = (obj.arrival_at_depot - obj.dj_departure_time).days
        row = [
            obj.create_date, obj.month_name, obj.transporter, obj.load_number, obj.driver_name, obj.truck_number, obj.customer_name,
            obj.dj_departure_time, getattr(obj, 'clockin_time', ''), arrival_at_customer, arrival_at_depot, days_spent,
            getattr(obj, 'departure_time_from_customer', ''), obj.total_distance, obj.total_time, obj.delivery_time, obj.efficiency_score, obj.current_status
        ]
        writer.writerow(row)
    return response
# --- Export Excel Report ---
from .export_utils import export_excel_report
import pandas as pd
import pytz
from datetime import datetime, timedelta
import traceback
from django.contrib.auth.decorators import login_required

def make_naive(dt):
    """Force a datetime or pandas Timestamp to be naive (no tzinfo, always in local time)."""
    if dt is None or (hasattr(pd, 'isna') and pd.isna(dt)):
        return None
    try:
        # Handle pandas Timestamp
        if hasattr(dt, 'tz_localize') or hasattr(dt, 'tzinfo'):
            # If tz-aware, convert to UTC then remove tzinfo
            if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                # If pandas Timestamp
                if hasattr(dt, 'to_pydatetime'):
                    dt = dt.to_pydatetime()
                # Convert to UTC then remove tzinfo
                dt = dt.astimezone(pytz.UTC).replace(tzinfo=None)
                return dt
            # If pandas Timestamp but tz-naive
            if hasattr(dt, 'to_pydatetime'):
                return dt.to_pydatetime()
            return dt
        # If string, try to parse
        if isinstance(dt, str):
            try:
                return pd.to_datetime(dt, utc=True).tz_convert(None).to_pydatetime()
            except Exception:
                return pd.to_datetime(dt, errors='coerce')
        return dt
    except Exception:
        return dt
import pandas as pd
import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Avg, Count, Sum, Min, Max, Q
from django.utils import timezone
from django.db import transaction
from datetime import datetime, timedelta
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.offline import plot
import json
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

from .models import CSVUpload, TruckPerformanceData, ProductivitySummary
from .forms import CSVUploadForm, BulkUploadForm


def truck_tracking_view(request):
    """View for tracking truck progress similar to Jumia order tracking"""
    # Get search query
    search_query = request.GET.get('search', '')
    
    # Base queryset
    all_trucks = TruckPerformanceData.objects.all()
    
    # Apply search filter if provided
    if search_query:
        all_trucks = all_trucks.filter(
            Q(load_number__icontains=search_query) |
            Q(driver_name__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(truck_id__icontains=search_query)
        )
    
    # Separate active and completed trucks
    active_trucks = all_trucks.exclude(current_status='journey_completed').order_by('-dj_departure_time')
    completed_trucks = all_trucks.filter(current_status='journey_completed').order_by('-arrival_at_depot')[:10]
    
    # Calculate progress for each truck
    for truck in active_trucks:
        truck.progress_percentage = truck.calculate_progress_percentage()
        truck.status_display = truck.get_current_status_display()
    
    for truck in completed_trucks:
        truck.progress_percentage = 100
        truck.status_display = truck.get_current_status_display()
    
    context = {
        'active_trucks': active_trucks,
        'completed_trucks': completed_trucks,
        'search_query': search_query,
        'total_trucks': all_trucks.count(),
    }
    
    return render(request, 'dashboard/truck_tracking.html', context)


def truck_detail_tracking(request, truck_id):
    """Detailed tracking view for a specific truck"""
    truck = get_object_or_404(TruckPerformanceData, id=truck_id)
    
    # Get progress steps
    progress_steps = truck.get_progress_steps()
    progress_percentage = truck.get_progress_percentage()
    
    context = {
        'truck': truck,
        'progress_steps': progress_steps,
        'progress_percentage': progress_percentage,
    }
    
    return render(request, 'dashboard/truck_detail_tracking.html', context)


def dashboard_view(request):
    """Main dashboard view with summary statistics and charts"""

    # Load number search
    load_search = request.GET.get('load_search', '').strip()

    # 1. Get all depot departures as the base
    depot_departures = TruckPerformanceData.objects.filter(csv_upload__upload_type='depot_departures')
    if load_search:
        depot_departures = depot_departures.filter(load_number__icontains=load_search)
    depot_departures = depot_departures.order_by('load_number', '-create_date')

    # 2. For each depot departure, merge/fill data from other files
    journeys_by_load = []
    for base in depot_departures:
        # Find all related records for this load_number (and truck_number) from all file types
        related = TruckPerformanceData.objects.filter(load_number=base.load_number, truck_number=base.truck_number)
        # Build a merged journey dict, starting from base
        merged = base
        # Fill missing fields from related records (prefer latest by create_date)
        for rel in related.order_by('-create_date'):
            for field in [
                'arrival_at_customer', 'service_time_at_customer', 'ave_arrival_time',
                'd1', 'd2', 'd3', 'd4', 'arrival_at_depot', 'customer_name',
                'planned_departure_time', 'departure_deviation_min', 'ave_departure',
                'comment_ave_tir', 'current_status', 'efficiency_score', 'total_distance',
                'total_time', 'delivery_time', 'mode_of_capture', 'driver_name', 'truck_number',
            ]:
                base_val = getattr(merged, field, None)
                rel_val = getattr(rel, field, None)
                if (not base_val or str(base_val) in ['None', 'Unknown', 'Unknown Customer', 'Unknown Driver', 'Unknown Vehicle', 'nan', '']) and rel_val and str(rel_val) not in ['None', 'Unknown', 'Unknown Customer', 'Unknown Driver', 'Unknown Vehicle', 'nan', '']:
                    setattr(merged, field, rel_val)
        # Map driver to vehicle name (if available)
        merged.driver_vehicle = f"{merged.driver_name} ({merged.truck_number})"
        # Calculate days spent in journey (if both departure and arrival at depot exist)
        if merged.dj_departure_time and merged.arrival_at_depot:
            days_spent = (merged.arrival_at_depot - merged.dj_departure_time).days
        else:
            days_spent = ''
        merged.days_spent = days_spent
        # Filter out unwanted rows:
        # - Unknown Driver and TRUCK_999
        # - Unknown Customer
        # - Status Pending Departure
        if (
            (merged.driver_name and merged.driver_name.strip() == 'Unknown Driver' and merged.truck_number and str(merged.truck_number).strip() == 'TRUCK_999') or
            (merged.customer_name and merged.customer_name.strip() == 'Unknown Customer') or
            (merged.current_status and merged.current_status.strip() == 'Pending Departure')
        ):
            continue
        journeys_by_load.append(merged)

    # Get summary statistics
    total_loads = TruckPerformanceData.objects.count()
    total_trucks = TruckPerformanceData.objects.values('truck_number').distinct().count()
    total_drivers = TruckPerformanceData.objects.values('driver_name').distinct().count()
    total_customers = TruckPerformanceData.objects.values('customer_name').distinct().count()

    # Get recent uploads
    recent_uploads = CSVUpload.objects.order_by('-uploaded_at')[:5]

    # Get recent performance data for display - prioritize records with real data
    recent_data = TruckPerformanceData.objects.exclude(
        driver_name='Unknown Driver'
    ).exclude(
        driver_name__isnull=True
    ).order_by('-created_at')[:10]

    # Calculate average efficiency score
    avg_efficiency = TruckPerformanceData.objects.aggregate(
        avg_score=Avg('efficiency_score')
    )['avg_score'] or 0

    # Get monthly performance data
    monthly_data = TruckPerformanceData.objects.values('month_name').annotate(
        total_loads=Count('id'),
        avg_efficiency=Avg('efficiency_score'),
        total_distance=Sum('total_distance')
    ).order_by('create_date')

    # Create performance charts
    charts = create_performance_charts()

    context = {
        'total_loads': total_loads,
        'total_trucks': total_trucks,
        'total_drivers': total_drivers,
        'total_customers': total_customers,
        'avg_efficiency': round(avg_efficiency, 2) if avg_efficiency else 0,
        'recent_uploads': recent_uploads,
        'recent_data': recent_data,
        'monthly_data': monthly_data,
        'charts': charts,
        'journeys_by_load': journeys_by_load,
    }

    return render(request, 'dashboard/dashboard.html', context)





def bulk_upload(request):
    """Handle bulk upload of all 6 CSV files"""
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        
        # Debug: Check what files are being received
        print("=== BULK UPLOAD DEBUG ===")
        print("Form is valid:", form.is_valid())
        print("Request FILES:", list(request.FILES.keys()))
        print("Form cleaned_data:", form.cleaned_data if form.is_valid() else "Form invalid")
        if not form.is_valid():
            print("Form errors:", form.errors)
        
        if form.is_valid():
            upload_count = 0
            error_count = 0
            
            # Define the file types and their corresponding form fields
            file_types = [
                ('depot_departures', 'depot_departures_file'),
                ('customer_timestamps', 'customer_timestamps_file'),
                ('distance_info', 'distance_info_file'),
                ('timestamps_duration', 'timestamps_duration_file'),
                ('avg_time_route', 'avg_time_route_file'),
                ('time_route_info', 'time_route_info_file'),
            ]
            
            for upload_type, field_name in file_types:
                uploaded_file = form.cleaned_data.get(field_name)
                print(f"Processing {field_name}: {uploaded_file}")
                
                if uploaded_file:
                    try:
                        # Create CSVUpload record
                        csv_upload = CSVUpload.objects.create(
                            name=uploaded_file.name,
                            upload_type=upload_type,
                            file=uploaded_file
                        )
                        print(f"Created CSVUpload record: {csv_upload.id}")
                        
                        # Process the file
                        if process_csv_file(csv_upload):
                            upload_count += 1
                            csv_upload.processed = True
                            csv_upload.save()
                            print(f"Successfully processed {uploaded_file.name}")
                        else:
                            error_count += 1
                            messages.error(request, f'Error processing {uploaded_file.name}')
                            print(f"Failed to process {uploaded_file.name}")
                            
                    except Exception as err:
                        error_count += 1
                        messages.error(request, f'Error with {uploaded_file.name}: {str(err)}')
                        print(f"Exception processing {uploaded_file.name}: {str(err)}")
            
            print(f"Upload summary: {upload_count} successful, {error_count} errors")
            
            if upload_count > 0:
                messages.success(request, f'Successfully processed {upload_count} files! You can now download the combined Excel report.')
                return redirect('dashboard:reports')
            if error_count > 0:
                messages.warning(request, f'{error_count} files had processing errors.')
            
            return redirect('dashboard:dashboard')
        else:
            # Form is not valid, show errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = BulkUploadForm()
    
    return render(request, 'dashboard/bulk_upload.html', {'form': form})

def get_fuzzy(row, keys, default=None):
    """Helper to get value from row using multiple possible keys"""
    # Try exact match first
    for key in keys:
        if key in row and pd.notna(row[key]) and str(row[key]).strip() != '':
            return row[key]
    # Try case-insensitive match
    row_keys_lower = {k.lower().strip(): k for k in row.index}
    for key in keys:
        k_lower = key.lower().strip()
        if k_lower in row_keys_lower:
            actual_key = row_keys_lower[k_lower]
            val = row[actual_key]
            if pd.notna(val) and str(val).strip() != '':
                return val
    return default


def extract_unified_truck_data(row, file_type):
    """Extract truck data from any file type and return standardized format"""
    data = {}
    
    if file_type == 'depot_departures':
        # File 1: Depot Departures Information
        dj_departure_time = pd.to_datetime(row.get('DJ Departure Time'), errors='coerce')
        dj_departure_time = make_naive(dj_departure_time)
        
        load_num = get_fuzzy(row, ['Load Number', 'Load Name', 'Load', 'Order No'], 'Unknown')
        truck_num = get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle', 'Truck'], 'Unknown')
        
        data = {
            'month_name': pd.to_datetime(row.get('Schedule Date', '2025-01-01')).strftime('%B'),
            'transporter': row.get('Depot', 'Unknown'),
            'load_number': str(load_num),
            'mode_of_capture': 'DJ',
            'driver_name': str(row.get('Driver Name', 'Unknown')),
            'truck_number': str(truck_num),
            'customer_name': 'Unknown Customer',  # Not in this file type
            'dj_departure_time': dj_departure_time,
            'departure_deviation_min': pd.to_numeric(row.get('Departure Time Difference (DJ vs Planned)'), errors='coerce'),
            'tlp_vol_hl': pd.to_numeric(row.get('TLP Vol HL', row.get('Tlp Vol Hl', row.get('Volume', 0))), errors='coerce'),
            'planned_arrival_time': make_naive(pd.to_datetime(row.get('Planned Arrival Time'), errors='coerce')),
        }

    elif file_type == 'customer_timestamps':
        # File 2: Customer Timestamps - TIME DATA IN MINUTES
        arrival_at_customer = pd.to_datetime(row.get('ArrivedAtCustomer(Odo)'), errors='coerce')
        arrival_at_customer = make_naive(arrival_at_customer)
        
        load_num = get_fuzzy(row, ['Load Number', 'Load Name', 'Load', 'load_name'], 'Unknown')
        truck_num = get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle'], 'Unknown')
        
        data = {
            'month_name': pd.to_datetime(row.get('schedule_date', '2025-01-01')).strftime('%B'),
            'transporter': row.get('Depot', 'Unknown'),
            'load_number': str(load_num),
            'mode_of_capture': 'DJ',
            'driver_name': str(row.get('DriverName', 'Unknown')),
            'truck_number': str(truck_num),
            'customer_name': str(row.get('customer_name', 'Unknown')),
            'arrival_at_customer': arrival_at_customer,
            'service_time_at_customer': pd.to_numeric(row.get('Total Time Spent @ Customer'), errors='coerce'),
            # These are actually time values in minutes, not distances
            'ave_arrival_time': pd.to_numeric(row.get('Total Time Spent @ Customer'), errors='coerce'),
            'd1': pd.to_numeric(row.get('Customer Gate To Offloading'), errors='coerce'),
            'd2': pd.to_numeric(row.get('Offloading to Invoice Completion'), errors='coerce'),
        }

    elif file_type == 'distance_info':
        # File 3: Distance Information - ACTUAL DISTANCE DATA
        planned_load_distance = pd.to_numeric(row.get('Planned Load Distance'), errors='coerce')
        planned_distance_to_customer = pd.to_numeric(row.get('PlannedDistanceToCustomer'), errors='coerce')
        km_deviation = pd.to_numeric(row.get('Load Distance Difference (Planned vs. DJ)'), errors='coerce')
        
        load_num = get_fuzzy(row, ['Load Number', 'Load Name', 'Load'], 'Unknown')
        truck_num = get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle'], 'Unknown')
        
        data = {
            'month_name': pd.to_datetime(row.get('Schedule Date', '2025-01-01')).strftime('%B'),
            'transporter': row.get('Depot', 'Unknown'),
            'load_number': str(load_num),
            'mode_of_capture': 'DJ',
            'driver_name': str(row.get('Driver Name', 'Unknown')),
            'truck_number': str(truck_num),
            'customer_name': str(row.get('Customer', 'Unknown')),
            'budgeted_kms': planned_load_distance,
            'PlannedDistanceToCustomer': planned_distance_to_customer,
            'km_deviation': km_deviation,
            'comment': '',
            # Keep other fields as before if needed
            'd1': planned_distance_to_customer,
            'd4': km_deviation,
        }

    elif file_type == 'timestamps_duration':
        # File 4: Timestamps and Duration
        arrival_at_depot = pd.to_datetime(row.get('ArriveAtDepot(Odo)'), errors='coerce')
        arrival_at_depot = make_naive(arrival_at_depot)
        data = {
            'month_name': pd.to_datetime(row.get('schedule_date', '2025-01-01')).strftime('%B'),
            'transporter': row.get('Depot', 'Unknown'),
            'load_number': str(row.get('load_name', 'Unknown')),
            'mode_of_capture': 'DJ',
            'driver_name': 'Unknown Driver',  # Not available in this file
            'truck_number': 'Unknown Vehicle',  # Not available in this file
            'customer_name': 'Unknown Customer',  # Not available in this file
            'arrival_at_depot': arrival_at_depot,
            'd1': pd.to_numeric(row.get('Load Start to Gate Exit'), errors='coerce'),
            'd2': pd.to_numeric(row.get('Depot Arrival to Gate Entry Complete'), errors='coerce'),
            'd3': pd.to_numeric(row.get('Gate Entry to load Completion'), errors='coerce'),
            'clock_out': make_naive(pd.to_datetime(row.get('LoadCompleted', row.get('Load Completed Time', row.get('Closure Time'))), errors='coerce')),
        }

    elif file_type == 'avg_time_route':
        # File 5: Average Time in Route
        data = {
            'month_name': 'Unknown',
            'transporter': 'Unknown',
            'load_number': 'Unknown',
            'mode_of_capture': 'DJ',
            'driver_name': 'Unknown Driver',
            'truck_number': 'Unknown Vehicle',
            'customer_name': str(row.get('customer_name', 'Unknown')),
            'comment_ave_tir': f"Average time difference: {row.get('Time In Route Difference ( DJ - Planned) (AVG)', 0)}",
        }

    elif file_type == 'time_route_info':
        # File 6: Time in Route Information
        data = {
            'month_name': pd.to_datetime(row.get('Schedule Date', '2025-01-01')).strftime('%B'),
            'transporter': row.get('Depot Code', 'Unknown'),
            'load_number': str(row.get('Load', 'Unknown')),
            'mode_of_capture': 'DJ',
            'driver_name': str(row.get('Driver', 'Unknown')),
            'truck_number': 'Unknown Vehicle',  # Not available in this file
            'customer_name': str(row.get('Customer', 'Unknown')),
            'd1': pd.to_numeric(row.get('Time In Route Difference ( DJ - Planned)'), errors='coerce'),
            'comment_ave_tir': f"Time in route: {row.get('Time in Route (min)', 0)} min, Planned: {row.get('Planned Time in Route (min)', 0)} min",
        }

    # Add common fields that are required
    if 'create_date' not in data:
        create_date = pd.to_datetime(data.get('month_name', 'January') + ' 2025', format='%B %Y', errors='coerce')
        data['create_date'] = make_naive(create_date)

    # Force all datetime fields in data to be naive
    for k, v in data.items():
        if isinstance(v, (pd.Timestamp, datetime)):
            data[k] = make_naive(v)

    return data


def process_csv_file(csv_upload):
    """Process uploaded CSV file based on its type and create TruckPerformanceData records"""
    try:
        # Read the CSV file
        file_path = csv_upload.file.path
        df = pd.read_csv(file_path)
        
        # Clean column names (remove extra spaces, standardize)
        df.columns = df.columns.str.strip()
        
        # Process based on upload type
        if csv_upload.upload_type == 'depot_departures':
            return process_depot_departures(df, csv_upload)
        elif csv_upload.upload_type == 'customer_timestamps':
            return process_customer_timestamps(df, csv_upload)
        elif csv_upload.upload_type == 'distance_info':
            return process_distance_info(df, csv_upload)
        elif csv_upload.upload_type == 'timestamps_duration':
            return process_timestamps_duration(df, csv_upload)
        elif csv_upload.upload_type == 'avg_time_route':
            return process_avg_time_route(df, csv_upload)
        elif csv_upload.upload_type == 'time_route_info':
            return process_time_route_info(df, csv_upload)
        else:
            return process_generic_csv(df, csv_upload)
            
    except Exception as err:
        print(f"Error processing CSV file: {str(err)}")
        return False


def process_depot_departures(df, csv_upload):
    """Process depot departures CSV file - File Type 1"""
    try:
        # Build a mapping of driver name to vehicle reg
        driver_vehicle_map = {}
        for index, row in df.iterrows():
            vehicle_reg = str(row.get('Vehicle Reg', '')).strip()
            driver_name = str(row.get('Driver Name', '')).strip()
            # Skip rows where Vehicle Reg is missing or 'Unknown'
            if not vehicle_reg or vehicle_reg.lower() == 'unknown' or not driver_name or driver_name.lower() == 'nan':
                continue
            driver_vehicle_map[driver_name] = vehicle_reg

        with transaction.atomic():
            for index, row in df.iterrows():
                # Extract data using unified function
                unified_data = extract_unified_truck_data(row, 'depot_departures')
                unified_data['csv_upload'] = csv_upload

                # Attach correct vehicle reg to driver name if possible
                driver_name = str(row.get('Driver Name', '')).strip()
                if driver_name in driver_vehicle_map:
                    unified_data['truck_number'] = driver_vehicle_map[driver_name]

                # Always make all datetime fields naive
                for dt_field in ['dj_departure_time', 'planned_departure_time', 'arrival_at_depot', 'create_date']:
                    if dt_field in unified_data:
                        unified_data[dt_field] = make_naive(unified_data[dt_field])

                # Add Planned Departure Time from depot departures file if present
                planned_departure_time = row.get('Planned Departure Time') or row.get('PlannedDepartureTime')
                if planned_departure_time is not None and planned_departure_time != '':
                    try:
                        from pandas import to_datetime
                        parsed = to_datetime(str(planned_departure_time), errors='coerce')
                        if pd.isna(parsed):
                            unified_data['planned_departure_time'] = None
                        else:
                            unified_data['planned_departure_time'] = make_naive(parsed)
                    except Exception as ex:
                        unified_data['planned_departure_time'] = None
                else:
                    pass

                # Clean up NaN for integer fields (convert to None)
                for int_field in ['departure_deviation_min', 'ave_departure']:
                    if int_field in unified_data:
                        val = unified_data[int_field]
                        if pd.isna(val):
                            unified_data[int_field] = None

                # Create or update the record - preserve existing good data
                existing_record, created = TruckPerformanceData.objects.get_or_create(
                    load_number=unified_data['load_number'],
                    create_date=unified_data['create_date'],
                    defaults=unified_data
                )

                # If record already exists, only update fields with real data
                if not created:
                    for field, value in unified_data.items():
                        if field not in ['load_number', 'create_date'] and value and str(value) not in ['Unknown', 'Unknown Customer', 'Unknown Driver', 'Unknown Vehicle']:
                            # Always make datetime fields naive on update
                            if isinstance(value, (datetime, pd.Timestamp)):
                                value = make_naive(value)
                            # Clean up NaN for integer fields on update
                            if field in ['departure_deviation_min', 'ave_departure'] and pd.isna(value):
                                value = None
                            setattr(existing_record, field, value)
                    existing_record.save()

        csv_upload.processed = True
        csv_upload.save()
        return True

    except Exception as err:
        print(f"Error processing depot departures: {str(err)}")
        return False


def process_customer_timestamps(df, csv_upload):
    """Process customer timestamps CSV file - File Type 2"""
    error_rows = []
    # Build a mapping from (load_number, create_date) to truck_number from depot_departures
    depot_map = {}
    from dashboard.models import TruckPerformanceData
    depot_qs = TruckPerformanceData.objects.filter(csv_upload__upload_type='depot_departures')
    for rec in depot_qs:
        key = (rec.load_number, rec.create_date)
        depot_map[key] = rec.truck_number

    with transaction.atomic():
        for index, row in df.iterrows():
            try:
                # Extract data using unified function
                unified_data = extract_unified_truck_data(row, 'customer_timestamps')
                unified_data['csv_upload'] = csv_upload

                # Always make all datetime fields naive
                for dt_field in ['arrival_at_customer', 'create_date']:
                    if dt_field in unified_data:
                        unified_data[dt_field] = make_naive(unified_data[dt_field])

                # Overwrite truck_number with exact Vehicle Reg from depot_departures if available
                key = (unified_data['load_number'], unified_data['create_date'])
                if key in depot_map:
                    unified_data['truck_number'] = depot_map[key]

                # Try to match with existing load, date, and truck number or create new record - preserve good data
                existing_record, created = TruckPerformanceData.objects.get_or_create(
                    load_number=unified_data['load_number'],
                    create_date=unified_data['create_date'],
                    truck_number=unified_data['truck_number'],
                    defaults=unified_data
                )

                # If record already exists, only update fields with real data
                if not created:
                    for field, value in unified_data.items():
                        if field not in ['load_number'] and value and str(value) not in ['Unknown', 'Unknown Customer', 'Unknown Driver', 'Unknown Vehicle']:
                            if isinstance(value, (datetime, pd.Timestamp)):
                                value = make_naive(value)
                            setattr(existing_record, field, value)
                    existing_record.save()
            except Exception as err:
                error_rows.append(f"Row {index+1}: {str(err)}")

    if error_rows:
        print("\n--- Error(s) processing customer timestamps ---")
        for err in error_rows:
            print(err)
        print("--- End error report ---\n")
        return False

    csv_upload.processed = True
    csv_upload.save()
    return True


def process_distance_info(df, csv_upload):
    """Process distance information CSV file"""
    try:
        current_rows = []
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Extract data using unified function with CORRECT file type
                    unified_data = extract_unified_truck_data(row, 'distance_info')
                    unified_data['csv_upload'] = csv_upload

                    # Always make all datetime fields naive
                    for dt_field in ['dj_departure_time', 'planned_departure_time', 'arrival_at_depot', 'create_date']:
                        if dt_field in unified_data:
                            unified_data[dt_field] = make_naive(unified_data[dt_field])

                    # Save the record
                    # We need to be careful not to overwrite everything if we only have distance info
                    # So we use update_or_create but we should probably try to find the existing one first
                    
                    # Check if Load Name exists in unified_data (it should)
                    load_number = unified_data.get('load_number')
                    if not load_number or load_number == 'Unknown':
                        continue

                    # Find existing record or create new
                    # Note: creating new might be risky if we don't have enough info, but distance info usually has Load Name
                    TruckPerformanceData.objects.update_or_create(
                        load_number=load_number,
                        defaults=unified_data
                    )
                except Exception as row_e:
                    print(f"Error processing row {index} in distance info: {row_e}")
                    traceback.print_exc()
        
        csv_upload.processed = True
        csv_upload.save()
        return True
    except Exception as err:
        print(f"Error in distance info processing: {err}")
        traceback.print_exc()
        return False


def process_timestamps_duration(df, csv_upload):
    """Process timestamps and duration CSV file"""
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                load_number = str(get_fuzzy(row, ['Load Number', 'Load Name', 'Load'], 'Unknown'))
                truck_number = str(get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle'], 'Unknown'))
                driver_name = str(get_fuzzy(row, ['Driver Name', 'DriverName', 'Driver'], 'Unknown Driver'))
                customer_name = str(get_fuzzy(row, ['Customer Name', 'customer_name', 'Customer'], 'Unknown Customer'))
                
                data = {
                    'csv_upload': csv_upload,
                    'create_date': make_naive(pd.to_datetime(row.get('Date', datetime.now().date()))),
                    'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                    'transporter': row.get('Transporter', row.get('Depot', 'Unknown')),
                    'load_number': load_number,
                    'mode_of_capture': 'DJ',
                    'driver_name': driver_name,
                    'truck_number': truck_number,
                    'customer_name': customer_name,
                    'dj_departure_time': make_naive(pd.to_datetime(row.get('Departure Time'), errors='coerce')),
                    'arrival_at_depot': make_naive(pd.to_datetime(row.get('Arrival Time'), errors='coerce')),
                    'clock_out': make_naive(pd.to_datetime(row.get('LoadCompleted'), errors='coerce')),
                    'comment_ave_tir': row.get('Duration Notes', ''),
                }
                # Make all datetime fields naive
                for dt_field in ['create_date', 'dj_departure_time', 'arrival_at_depot']:
                    if dt_field in data:
                        data[dt_field] = make_naive(data[dt_field])
                TruckPerformanceData.objects.update_or_create(
                    load_number=load_number,
                    truck_number=truck_number,
                    defaults=data
                )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as err:
        print(f"Error processing timestamps duration: {str(err)}")
        return False


def process_avg_time_route(df, csv_upload):
    """Process average time in route CSV file"""
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                load_number = str(get_fuzzy(row, ['Load Number', 'Load Name', 'Load'], 'Unknown'))
                truck_number = str(get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle'], 'Unknown'))
                driver_name = str(get_fuzzy(row, ['Driver Name', 'DriverName', 'Driver'], 'Unknown Driver'))
                customer_name = str(get_fuzzy(row, ['Customer Name', 'customer_name', 'Customer'], 'Unknown Customer'))
                
                data = {
                    'csv_upload': csv_upload,
                    'create_date': make_naive(pd.to_datetime(row.get('Date', datetime.now().date()))),
                    'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                    'transporter': row.get('Transporter', row.get('Depot', 'Unknown')),
                    'load_number': load_number,
                    'mode_of_capture': 'Average Time',
                    'driver_name': row.get('Driver Name', 'Unknown Driver'),
                    'truck_number': truck_number,
                    'customer_name': row.get('Customer Name', 'Unknown Customer'),
                    'ave_arrival_time': make_naive(pd.to_datetime(row.get('Average Arrival Time'), errors='coerce')),
                    'comment_ave_tir': row.get('Time Comments', ''),
                }
                # Make all datetime fields naive
                for dt_field in ['create_date', 'ave_arrival_time']:
                    if dt_field in data:
                        data[dt_field] = make_naive(data[dt_field])
                TruckPerformanceData.objects.update_or_create(
                    load_number=load_number,
                    truck_number=truck_number,
                    defaults=data
                )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as err:
        print(f"Error processing avg time route: {str(err)}")
        return False


def process_time_route_info(df, csv_upload):
    """Process time in route information CSV file"""
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                load_number = str(get_fuzzy(row, ['Load Number', 'Load Name', 'Load'], 'Unknown'))
                truck_number = str(get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle'], 'Unknown'))
                driver_name = str(get_fuzzy(row, ['Driver Name', 'DriverName', 'Driver'], 'Unknown Driver'))
                customer_name = str(get_fuzzy(row, ['Customer Name', 'customer_name', 'Customer'], 'Unknown Customer'))
                
                data = {
                    'csv_upload': csv_upload,
                    'create_date': make_naive(pd.to_datetime(row.get('Date', datetime.now().date()))),
                    'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                    'transporter': row.get('Transporter', row.get('Depot', 'Unknown')),
                    'load_number': load_number,
                    'mode_of_capture': 'DJ',
                    'driver_name': driver_name,
                    'truck_number': truck_number,
                    'customer_name': customer_name,
                    'dj_departure_time': make_naive(pd.to_datetime(row.get('Route Start Time'), errors='coerce')),
                    'arrival_at_depot': make_naive(pd.to_datetime(row.get('Route End Time'), errors='coerce')),
                    'comment_ave_tir': row.get('Route Comments', ''),
                }
                # Make all datetime fields naive
                for dt_field in ['create_date', 'dj_departure_time', 'arrival_at_depot']:
                    if dt_field in data:
                        data[dt_field] = make_naive(data[dt_field])
                TruckPerformanceData.objects.update_or_create(
                    load_number=load_number,
                    truck_number=truck_number,
                    defaults=data
                )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as err:
        print(f"Error processing time route info: {str(err)}")
        return False


def process_generic_csv(df, csv_upload):
    """Process generic CSV file with best-effort field mapping"""
    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                # Try to extract common fields - prioritize actual column names from uploads
                load_number = str(get_fuzzy(row, ['Load Number', 'Load Name', 'Load Name 1', 'Load', 'ID'], 'Unknown'))
                truck_number = str(get_fuzzy(row, ['Vehicle Reg', 'Truck Number', 'Vehicle', 'Truck'], 'Unknown'))
                driver_name = str(get_fuzzy(row, ['Driver Name', 'DriverName', 'Driver'], 'Unknown Driver'))
                customer_name = str(get_fuzzy(row, ['Customer Name', 'customer_name', 'Customer'], 'Unknown Customer'))
                
                data = {
                    'csv_upload': csv_upload,
                    'create_date': pd.to_datetime(row.get('schedule_date', row.get('Create Date', row.get('Date', datetime.now().date())))).date(),
                    'month_name': pd.to_datetime(row.get('schedule_date', row.get('Create Date', row.get('Date', datetime.now())))).strftime('%B'),
                    'transporter': row.get('Transporter', row.get('Depot', row.get('Company', 'Unknown'))),
                    'load_number': load_number,
                    'mode_of_capture': 'DJ',
                    'driver_name': driver_name,
                    'truck_number': truck_number,
                    'customer_name': customer_name,
                    'comment_ave_tir': str(row.get('Comments', row.get('Notes', ''))),
                }
                
                TruckPerformanceData.objects.update_or_create(
                    load_number=load_number,
                    truck_number=truck_number,
                    defaults=data
                )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as err:
        print(f"Error processing generic CSV: {str(err)}")
        return False


def create_performance_charts():
    """Create interactive charts for the dashboard"""
    charts = {}
    
    try:
        # Get data for charts
        performance_data = TruckPerformanceData.objects.all()
        
        if performance_data.exists():
            # Monthly performance chart
            monthly_stats = performance_data.values('month_name').annotate(
                total_loads=Count('id'),
                avg_efficiency=Avg('efficiency_score'),
                total_distance=Sum('total_distance')
            ).order_by('create_date')
            
            if monthly_stats:
                months = [item['month_name'] for item in monthly_stats]
                loads = [item['total_loads'] for item in monthly_stats]
                
                fig = go.Figure(data=[
                    go.Bar(name='Total Loads', x=months, y=loads)
                ])
                fig.update_layout(title='Monthly Load Performance', xaxis_title='Month', yaxis_title='Number of Loads')
                charts['monthly_performance'] = plot(fig, output_type='div', include_plotlyjs=False)
            
            # Transporter performance chart
            transporter_stats = performance_data.values('transporter').annotate(
                total_loads=Count('id'),
                avg_efficiency=Avg('efficiency_score')
            ).order_by('-total_loads')[:10]
            
            if transporter_stats:
                transporters = [item['transporter'] for item in transporter_stats]
                loads = [item['total_loads'] for item in transporter_stats]
                
                fig = go.Figure(data=[
                    go.Bar(name='Loads by Transporter', x=transporters, y=loads)
                ])
                fig.update_layout(title='Top Transporters by Load Count', xaxis_title='Transporter', yaxis_title='Total Loads')
                charts['transporter_performance'] = plot(fig, output_type='div', include_plotlyjs=False)
    
    except Exception as e:
        print(f"Error creating charts: {str(e)}")
    
    return charts




def create_executive_summary_sheet(ws):
    """Create executive summary sheet"""
    # Header
    ws['A1'] = 'Truck Productivity Dashboard - Executive Summary'
    ws['A1'].font = Font(bold=True, size=16)
    
    # Key metrics
    total_loads = TruckPerformanceData.objects.count()
    total_trucks = TruckPerformanceData.objects.values('truck_number').distinct().count()
    total_drivers = TruckPerformanceData.objects.values('driver_name').distinct().count()
    total_customers = TruckPerformanceData.objects.values('customer_name').distinct().count()
    
    avg_efficiency = TruckPerformanceData.objects.aggregate(avg=Avg('efficiency_score'))['avg'] or 0
    total_distance = TruckPerformanceData.objects.aggregate(total=Sum('total_distance'))['total'] or 0
    
    # Add metrics to sheet
    metrics = [
        ['Metric', 'Value'],
        ['Total Loads', total_loads],
        ['Total Trucks', total_trucks],
        ['Total Drivers', total_drivers],
        ['Total Customers', total_customers],
        ['Average Efficiency Score', round(avg_efficiency, 2)],
        ['Total Distance (km)', round(total_distance, 2)],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for row_num, row_data in enumerate(metrics, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 3:  # Header row
                cell.font = Font(bold=True)


def create_detailed_report_sheet(ws):
    """Create detailed report sheet with all truck performance data"""
    
    # Helper function to convert integer time to string format
    def format_time_from_int(time_int):
        if time_int is None:
            return ''
        try:
            # Assuming time_int is in minutes since midnight
            hours = time_int // 60
            minutes = time_int % 60
            return f"{hours:02d}:{minutes:02d}:00"
        except (TypeError, ValueError):
            return str(time_int) if time_int else ''
    
    # Headers matching the required attributes
    headers = [
        'Create Date', 'Month Name', 'Transporter', 'Load Number', 'Mode Of Capture',
        'Driver Name', 'Truck Number', 'Customer Name', 'Clock-In', 'DJ Departure Time',
        'Arrival At Depot', 'AVE Arrival Time',
        'D1', 'D2', 'D3', 'D4', 'Comment Ave TIR',
        'Total Distance', 'Total Time (hrs)', 'Efficiency (km/h)'
    ]
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Add data
    data = TruckPerformanceData.objects.all().order_by('-create_date')
    # Build a mapping from driver_name to truck_number from depot_departures (never Unknown)
    depot_departures = TruckPerformanceData.objects.filter(csv_upload__upload_type='depot_departures')
    driver_to_truck = dict(
        depot_departures.exclude(driver_name__in=['Unknown Driver', '', None])
        .exclude(truck_number__in=['Unknown', '', None])
        .values_list('driver_name', 'truck_number')
    )
    for row_num, item in enumerate(data, 2):
        clockin_str = item.clockin_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(item, 'clockin_time') and item.clockin_time else ''
        # Always map truck_number from depot_departures using driver_name if possible, never show Unknown
        mapped_truck = driver_to_truck.get(item.driver_name, '')
        if mapped_truck:
            truck_number = mapped_truck
        else:
            truck_number = ''
        row_data = [
            item.create_date.strftime('%Y-%m-%d') if item.create_date else '',
            item.month_name or '',
            item.transporter or '',
            item.load_number or '',
            'DJ',  # Set mode of capture to DJ as requested
            item.driver_name or '',
            truck_number,
            item.customer_name or '',
            clockin_str,
            item.dj_departure_time.strftime('%Y-%m-%d %H:%M:%S') if item.dj_departure_time else '',
            item.arrival_at_depot.strftime('%Y-%m-%d %H:%M:%S') if item.arrival_at_depot else '',
            format_time_from_int(item.ave_arrival_time),  # Handle integer time field
            item.d1 or '',
            item.d2 or '',
            item.d3 or '',
            item.d4 or '',
            item.comment_ave_tir or '',
            item.total_distance or '',
            f"{item.total_time:.2f}" if item.total_time is not None else '',
            f"{item.efficiency_score:.2f}" if item.efficiency_score is not None else '',
        ]
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_transporter_summary_sheet(ws):
    """Create transporter summary sheet"""
    # Headers
    headers = ['Transporter', 'Total Loads', 'Total Distance', 'Average Efficiency', 'Total Drivers']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by transporter
    transporter_data = TruckPerformanceData.objects.values('transporter').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score'),
        total_drivers=Count('driver_name', distinct=True)
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(transporter_data, 2):
        row_data = [
            item['transporter'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
            item['total_drivers'],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_customer_summary_sheet(ws):
    """Create customer summary sheet"""
    # Headers
    headers = ['Customer Name', 'Total Loads', 'Total Distance', 'Average Efficiency']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by customer
    customer_data = TruckPerformanceData.objects.values('customer_name').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score')
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(customer_data, 2):
        row_data = [
            item['customer_name'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_driver_performance_sheet(ws):
    """Create driver performance sheet"""
    # Headers
    headers = ['Driver Name', 'Truck Number', 'Total Loads', 'Total Distance', 'Average Efficiency', 'Total Customers']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by driver
    driver_data = TruckPerformanceData.objects.values('driver_name', 'truck_number').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score'),
        total_customers=Count('customer_name', distinct=True)
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(driver_data, 2):
        row_data = [
            item['driver_name'],
            item['truck_number'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
            item['total_customers'],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def reports_view(request):
    """Minimal reports view: only show export button/link."""
    return render(request, 'dashboard/reports.html')




def clear_all_data(request):
    """Clear all uploaded files and data for fresh upload"""
    if request.method == 'POST':
        try:
            # Delete all TruckPerformanceData records
            deleted_performance = TruckPerformanceData.objects.all().delete()
            
            # Delete all CSVUpload records
            deleted_uploads = CSVUpload.objects.all().delete()
            
            # Delete all ProductivitySummary records
            deleted_summaries = ProductivitySummary.objects.all().delete()
            
            messages.success(
                request, 
                f'Successfully cleared all data! Deleted {deleted_performance[0]} performance records, '
                f'{deleted_uploads[0]} upload records, and {deleted_summaries[0]} summary records.'
            )
            
        except Exception as e:
            messages.error(request, f'Error clearing data: {str(e)}')
    
    return redirect('dashboard:bulk_upload')


def truck_status_api(request):
    """API endpoint for real-time truck status updates"""
    search_query = request.GET.get('search', '')
    
    # Base queryset
    all_trucks = TruckPerformanceData.objects.all()
    
    # Apply search filter if provided
    if search_query:
        all_trucks = all_trucks.filter(
            Q(load_number__icontains=search_query) |
            Q(driver_name__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(truck_id__icontains=search_query)
        )
    
    # Separate active and completed trucks
    active_trucks = all_trucks.exclude(current_status='journey_completed')
    completed_trucks = all_trucks.filter(current_status='journey_completed')
    
    # Prepare data for JSON response
    active_data = []
    for truck in active_trucks:
        pass  # Placeholder or implement logic as needed
