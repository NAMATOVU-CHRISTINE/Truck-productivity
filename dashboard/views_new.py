from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Avg, Count, Sum, Min, Max
from django.utils import timezone
from datetime import datetime, timedelta
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.offline import plot
import json
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

from .models import CSVUpload, TruckPerformanceData, ProductivitySummary
from .forms import CSVUploadForm, BulkUploadForm


def dashboard_view(request):
    """Main dashboard view with summary statistics and charts"""
    # Get summary statistics
    total_loads = TruckPerformanceData.objects.count()
    total_trucks = TruckPerformanceData.objects.values('truck_number').distinct().count()
    total_drivers = TruckPerformanceData.objects.values('driver_name').distinct().count()
    total_customers = TruckPerformanceData.objects.values('customer_name').distinct().count()
    
    # Get recent uploads
    recent_uploads = CSVUpload.objects.order_by('-uploaded_at')[:5]
    
    # Calculate average efficiency score
    avg_efficiency = TruckPerformanceData.objects.aggregate(
        avg_score=Avg('efficiency_score')
    )['avg_score'] or 0
    
    # Get monthly performance data
    monthly_data = TruckPerformanceData.objects.values('month_name').annotate(
        total_loads=Count('id'),
        avg_efficiency=Avg('efficiency_score'),
        total_distance=Sum('total_distance')
    ).order_by('create_date')
    
    # Create performance charts
    charts = create_performance_charts()
    
    context = {
        'total_loads': total_loads,
        'total_trucks': total_trucks,
        'total_drivers': total_drivers,
        'total_customers': total_customers,
        'avg_efficiency': round(avg_efficiency, 2),
        'recent_uploads': recent_uploads,
        'monthly_data': monthly_data,
        'charts': charts,
    }
    
    return render(request, 'dashboard/dashboard.html', context)


def upload_csv(request):
    """Handle single CSV file upload"""
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_upload = form.save()
            
            # Process the uploaded file
            try:
                if process_csv_file(csv_upload):
                    messages.success(request, f'File "{csv_upload.name}" uploaded and processed successfully!')
                else:
                    messages.error(request, f'Error processing file "{csv_upload.name}". Please check the file format.')
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
            
            return redirect('dashboard:dashboard')
    else:
        form = CSVUploadForm()
    
    return render(request, 'dashboard/upload.html', {'form': form})


def bulk_upload(request):
    """Handle bulk upload of all 6 CSV files"""
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload_count = 0
            error_count = 0
            
            # Define the file types and their corresponding form fields
            file_types = [
                ('depot_departures', 'file_depot_departures'),
                ('customer_timestamps', 'file_customer_timestamps'),
                ('distance_info', 'file_distance_info'),
                ('timestamps_duration', 'file_timestamps_duration'),
                ('avg_time_route', 'file_avg_time_route'),
                ('time_route_info', 'file_time_route_info'),
            ]
            
            for upload_type, field_name in file_types:
                uploaded_file = request.FILES.get(field_name)
                if uploaded_file:
                    try:
                        # Create CSVUpload record
                        csv_upload = CSVUpload.objects.create(
                            name=uploaded_file.name,
                            upload_type=upload_type,
                            file=uploaded_file
                        )
                        
                        # Process the file
                        if process_csv_file(csv_upload):
                            upload_count += 1
                            csv_upload.processed = True
                            csv_upload.save()
                        else:
                            error_count += 1
                            messages.error(request, f'Error processing {uploaded_file.name}')
                            
                    except Exception as e:
                        error_count += 1
                        messages.error(request, f'Error with {uploaded_file.name}: {str(e)}')
            
            if upload_count > 0:
                messages.success(request, f'Successfully processed {upload_count} files!')
            if error_count > 0:
                messages.warning(request, f'{error_count} files had processing errors.')
            
            return redirect('dashboard:dashboard')
    else:
        form = BulkUploadForm()
    
    return render(request, 'dashboard/bulk_upload.html', {'form': form})


def process_csv_file(csv_upload):
    """Process uploaded CSV file based on its type and create TruckPerformanceData records"""
    try:
        # Read the CSV file
        file_path = csv_upload.file.path
        df = pd.read_csv(file_path)
        
        # Clean column names (remove extra spaces, standardize)
        df.columns = df.columns.str.strip()
        
        # Process based on upload type
        if csv_upload.upload_type == 'depot_departures':
            return process_depot_departures(df, csv_upload)
        elif csv_upload.upload_type == 'customer_timestamps':
            return process_customer_timestamps(df, csv_upload)
        elif csv_upload.upload_type == 'distance_info':
            return process_distance_info(df, csv_upload)
        elif csv_upload.upload_type == 'timestamps_duration':
            return process_timestamps_duration(df, csv_upload)
        elif csv_upload.upload_type == 'avg_time_route':
            return process_avg_time_route(df, csv_upload)
        elif csv_upload.upload_type == 'time_route_info':
            return process_time_route_info(df, csv_upload)
        else:
            return process_generic_csv(df, csv_upload)
            
    except Exception as e:
        print(f"Error processing CSV file: {str(e)}")
        return False


def process_depot_departures(df, csv_upload):
    """Process depot departures CSV file"""
    try:
        for index, row in df.iterrows():
            # Map CSV columns to model fields
            # Try multiple possible customer name columns
            customer_name = row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client') or 'Unknown Customer'
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Create Date', row.get('Date', datetime.now().date()))).date(),
                'month_name': pd.to_datetime(row.get('Create Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': str(row.get('Load Number', f'LOAD_{index}')),
                'mode_of_capture': row.get('Mode Of Capture', 'Manual'),
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': str(row.get('Truck Number', f'TRUCK_{index}')),
                'customer_name': customer_name,
                'dj_departure_time': pd.to_datetime(row.get('DJ Departure Time'), errors='coerce'),
                'arrival_at_depot': pd.to_datetime(row.get('Arrival At Depot'), errors='coerce'),
                'ave_arrival_time': pd.to_datetime(row.get('AVE Arrival Time'), errors='coerce'),
                'd1': pd.to_numeric(row.get('D1'), errors='coerce'),
                'd2': pd.to_numeric(row.get('D2'), errors='coerce'),
                'd3': pd.to_numeric(row.get('D3'), errors='coerce'),
                'd4': pd.to_numeric(row.get('D4'), errors='coerce'),
                'comment_ave_tir': row.get('Comment Ave TIR', ''),
            }
            
            # Create or update the record
            TruckPerformanceData.objects.update_or_create(
                load_number=data['load_number'],
                create_date=data['create_date'],
                truck_number=data['truck_number'],
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing depot departures: {str(e)}")
        return False


def process_customer_timestamps(df, csv_upload):
    """Process customer timestamps CSV file"""
    try:
        for index, row in df.iterrows():
            # Try to find existing record or create new one
            load_number = str(row.get('Load Number', f'LOAD_{index}'))
            truck_number = str(row.get('Truck Number', f'TRUCK_{index}'))
            
            # Try all possible customer name columns, including lowercase/underscore
            customer_name = (
                row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client')
                or row.get('customer_name') or row.get('customer') or row.get('client_name') or row.get('client') or ''
            )
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Date', datetime.now().date())).date(),
                'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': load_number,
                'mode_of_capture': row.get('Mode Of Capture', 'Timestamp'),
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': truck_number,
                'customer_name': customer_name,
                'ave_arrival_time': pd.to_datetime(row.get('Customer Arrival Time'), errors='coerce'),
                'comment_ave_tir': row.get('Notes', ''),
            }
            
            # Always update with latest customer_name and related fields
            obj, created = TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing customer timestamps: {str(e)}")
        return False


def process_distance_info(df, csv_upload):
    """Process distance information CSV file"""
    try:
        for index, row in df.iterrows():
            load_number = str(row.get('Load Number', f'LOAD_{index}'))
            truck_number = str(row.get('Truck Number', f'TRUCK_{index}'))
            
            customer_name = row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client') or ''
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Date', datetime.now().date())).date(),
                'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': load_number,
                'mode_of_capture': 'Distance Tracking',
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': truck_number,
                'customer_name': customer_name,
                'd1': pd.to_numeric(row.get('Distance 1', row.get('D1')), errors='coerce'),
                'd2': pd.to_numeric(row.get('Distance 2', row.get('D2')), errors='coerce'),
                'd3': pd.to_numeric(row.get('Distance 3', row.get('D3')), errors='coerce'),
                'd4': pd.to_numeric(row.get('Distance 4', row.get('D4')), errors='coerce'),
                'comment_ave_tir': row.get('Comments', ''),
            }
            
            TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing distance info: {str(e)}")
        return False


def process_timestamps_duration(df, csv_upload):
    """Process timestamps and duration CSV file"""
    try:
        for index, row in df.iterrows():
            load_number = str(row.get('Load Number', f'LOAD_{index}'))
            truck_number = str(row.get('Truck Number', f'TRUCK_{index}'))
            
            customer_name = row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client') or ''
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Date', datetime.now().date())).date(),
                'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': load_number,
                'mode_of_capture': 'Duration Tracking',
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': truck_number,
                'customer_name': customer_name,
                'dj_departure_time': pd.to_datetime(row.get('Departure Time'), errors='coerce'),
                'arrival_at_depot': pd.to_datetime(row.get('Arrival Time'), errors='coerce'),
                'comment_ave_tir': row.get('Duration Notes', ''),
            }
            
            TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing timestamps duration: {str(e)}")
        return False


def process_avg_time_route(df, csv_upload):
    """Process average time in route CSV file"""
    try:
        for index, row in df.iterrows():
            load_number = str(row.get('Load Number', f'LOAD_{index}'))
            truck_number = str(row.get('Truck Number', f'TRUCK_{index}'))
            
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Date', datetime.now().date())).date(),
                'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': load_number,
                'mode_of_capture': 'Average Time',
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': truck_number,
                'customer_name': row.get('Customer Name', ''),
                'ave_arrival_time': pd.to_datetime(row.get('Average Arrival Time'), errors='coerce'),
                'comment_ave_tir': row.get('Time Comments', ''),
            }
            
            TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing avg time route: {str(e)}")
        return False


def process_time_route_info(df, csv_upload):
    """Process time in route information CSV file"""
    try:
        for index, row in df.iterrows():
            load_number = str(row.get('Load Number', f'LOAD_{index}'))
            truck_number = str(row.get('Truck Number', f'TRUCK_{index}'))
            
            customer_name = row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client') or ''
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Date', datetime.now().date())).date(),
                'month_name': pd.to_datetime(row.get('Date', datetime.now())).strftime('%B'),
                'transporter': row.get('Transporter', 'Unknown'),
                'load_number': load_number,
                'mode_of_capture': 'Route Info',
                'driver_name': row.get('Driver Name', 'Unknown Driver'),
                'truck_number': truck_number,
                'customer_name': customer_name,
                'dj_departure_time': pd.to_datetime(row.get('Route Start Time'), errors='coerce'),
                'arrival_at_depot': pd.to_datetime(row.get('Route End Time'), errors='coerce'),
                'comment_ave_tir': row.get('Route Comments', ''),
            }
            
            TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing time route info: {str(e)}")
        return False


def process_generic_csv(df, csv_upload):
    """Process generic CSV file with best-effort field mapping"""
    try:
        for index, row in df.iterrows():
            # Try to extract common fields
            load_number = str(row.get('Load Number', row.get('Load', row.get('ID', f'LOAD_{index}'))))
            truck_number = str(row.get('Truck Number', row.get('Vehicle', row.get('Truck', f'TRUCK_{index}'))))
            
            customer_name = row.get('Customer Name') or row.get('Customer') or row.get('Client Name') or row.get('Client') or ''
            data = {
                'csv_upload': csv_upload,
                'create_date': pd.to_datetime(row.get('Create Date', row.get('Date', datetime.now().date()))).date(),
                'month_name': pd.to_datetime(row.get('Create Date', row.get('Date', datetime.now()))).strftime('%B'),
                'transporter': row.get('Transporter', row.get('Company', 'Unknown')),
                'load_number': load_number,
                'mode_of_capture': 'Generic Import',
                'driver_name': row.get('Driver Name', row.get('Driver', 'Unknown Driver')),
                'truck_number': truck_number,
                'customer_name': customer_name,
                'comment_ave_tir': str(row.get('Comments', row.get('Notes', ''))),
            }
            
            TruckPerformanceData.objects.update_or_create(
                load_number=load_number,
                truck_number=truck_number,
                defaults=data
            )
        
        csv_upload.processed = True
        csv_upload.save()
        return True
        
    except Exception as e:
        print(f"Error processing generic CSV: {str(e)}")
        return False


def create_performance_charts():
    """Create interactive charts for the dashboard"""
    charts = {}
    
    try:
        # Get data for charts
        performance_data = TruckPerformanceData.objects.all()
        
        if performance_data.exists():
            # Monthly performance chart
            monthly_stats = performance_data.values('month_name').annotate(
                total_loads=Count('id'),
                avg_efficiency=Avg('efficiency_score'),
                total_distance=Sum('total_distance')
            ).order_by('create_date')
            
            if monthly_stats:
                months = [item['month_name'] for item in monthly_stats]
                loads = [item['total_loads'] for item in monthly_stats]
                
                fig = go.Figure(data=[
                    go.Bar(name='Total Loads', x=months, y=loads)
                ])
                fig.update_layout(title='Monthly Load Performance', xaxis_title='Month', yaxis_title='Number of Loads')
                charts['monthly_performance'] = plot(fig, output_type='div', include_plotlyjs=False)
            
            # Transporter performance chart
            transporter_stats = performance_data.values('transporter').annotate(
                total_loads=Count('id'),
                avg_efficiency=Avg('efficiency_score')
            ).order_by('-total_loads')[:10]
            
            if transporter_stats:
                transporters = [item['transporter'] for item in transporter_stats]
                loads = [item['total_loads'] for item in transporter_stats]
                
                fig = go.Figure(data=[
                    go.Bar(name='Loads by Transporter', x=transporters, y=loads)
                ])
                fig.update_layout(title='Top Transporters by Load Count', xaxis_title='Transporter', yaxis_title='Total Loads')
                charts['transporter_performance'] = plot(fig, output_type='div', include_plotlyjs=False)
    
    except Exception as e:
        print(f"Error creating charts: {str(e)}")
    
    return charts


def export_report(request):
    """Export comprehensive Excel report with all data"""
    try:
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Executive Summary sheet
        ws_summary = wb.create_sheet("Executive Summary")
        create_executive_summary_sheet(ws_summary)
        
        # Create Detailed Report sheet
        ws_detailed = wb.create_sheet("Detailed Report")
        create_detailed_report_sheet(ws_detailed)
        
        # Create Transporter Summary sheet
        ws_transporter = wb.create_sheet("Transporter Summary")
        create_transporter_summary_sheet(ws_transporter)
        
        # Create Customer Summary sheet
        ws_customer = wb.create_sheet("Customer Summary")
        create_customer_summary_sheet(ws_customer)
        
        # Create Driver Performance sheet
        ws_driver = wb.create_sheet("Driver Performance")
        create_driver_performance_sheet(ws_driver)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Truck_Productivity_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating report: {str(e)}')
        return redirect('dashboard:dashboard')


def create_executive_summary_sheet(ws):
    """Create executive summary sheet"""
    # Header
    ws['A1'] = 'Truck Productivity Dashboard - Executive Summary'
    ws['A1'].font = Font(bold=True, size=16)
    
    # Key metrics
    total_loads = TruckPerformanceData.objects.count()
    total_trucks = TruckPerformanceData.objects.values('truck_number').distinct().count()
    total_drivers = TruckPerformanceData.objects.values('driver_name').distinct().count()
    total_customers = TruckPerformanceData.objects.values('customer_name').distinct().count()
    
    avg_efficiency = TruckPerformanceData.objects.aggregate(avg=Avg('efficiency_score'))['avg'] or 0
    total_distance = TruckPerformanceData.objects.aggregate(total=Sum('total_distance'))['total'] or 0
    
    # Add metrics to sheet
    metrics = [
        ['Metric', 'Value'],
        ['Total Loads', total_loads],
        ['Total Trucks', total_trucks],
        ['Total Drivers', total_drivers],
        ['Total Customers', total_customers],
        ['Average Efficiency Score', round(avg_efficiency, 2)],
        ['Total Distance (km)', round(total_distance, 2)],
        ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for row_num, row_data in enumerate(metrics, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 3:  # Header row
                cell.font = Font(bold=True)


def create_detailed_report_sheet(ws):
    """Create detailed report sheet with all truck performance data"""
    # Headers matching the required attributes
    headers = [
        'Create Date', 'Month Name', 'Transporter', 'Load Number', 'Mode Of Capture',
        'Driver Name', 'Truck Number', 'Customer Name', 'Clock In', 'DJ Departure Time',
        'Arrival At Depot', 'AVE Arrival Time', 'D1', 'D2', 'D3', 'D4', 'Comment Ave TIR',
        'Total Distance', 'Total Time', 'Efficiency Score'
    ]
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Add data
    data = TruckPerformanceData.objects.all().order_by('-create_date')
    for row_num, item in enumerate(data, 2):
        # Robustly calculate Clock In: DJ Departure Time minus 30 minutes
        clock_in_str = ''
        if item.dj_departure_time:
            try:
                clock_in_time = item.dj_departure_time - timedelta(minutes=30)
                clock_in_str = clock_in_time.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                clock_in_str = ''
        row_data = [
            item.create_date.strftime('%Y-%m-%d') if item.create_date else '',
            item.month_name or '',
            item.transporter or '',
            item.load_number or '',
            item.mode_of_capture or '',
            item.driver_name or '',
            item.truck_number or '',
            item.customer_name or '',
            clock_in_str,
            item.dj_departure_time.strftime('%Y-%m-%d %H:%M:%S') if item.dj_departure_time else '',
            item.arrival_at_depot.strftime('%Y-%m-%d %H:%M:%S') if item.arrival_at_depot else '',
            item.ave_arrival_time.strftime('%Y-%m-%d %H:%M:%S') if item.ave_arrival_time else '',
            item.d1 or '',
            item.d2 or '',
            item.d3 or '',
            item.d4 or '',
            item.comment_ave_tir or '',
            item.total_distance or '',
            item.total_time or '',
            item.efficiency_score or '',
        ]
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_transporter_summary_sheet(ws):
    """Create transporter summary sheet"""
    # Headers
    headers = ['Transporter', 'Total Loads', 'Total Distance', 'Average Efficiency', 'Total Drivers']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by transporter
    transporter_data = TruckPerformanceData.objects.values('transporter').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score'),
        total_drivers=Count('driver_name', distinct=True)
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(transporter_data, 2):
        row_data = [
            item['transporter'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
            item['total_drivers'],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_customer_summary_sheet(ws):
    """Create customer summary sheet"""
    # Headers
    headers = ['Customer Name', 'Total Loads', 'Total Distance', 'Average Efficiency']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by customer
    customer_data = TruckPerformanceData.objects.values('customer_name').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score')
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(customer_data, 2):
        row_data = [
            item['customer_name'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def create_driver_performance_sheet(ws):
    """Create driver performance sheet"""
    # Headers
    headers = ['Driver Name', 'Truck Number', 'Total Loads', 'Total Distance', 'Average Efficiency', 'Total Customers']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Aggregate data by driver
    driver_data = TruckPerformanceData.objects.values('driver_name', 'truck_number').annotate(
        total_loads=Count('id'),
        total_distance=Sum('total_distance'),
        avg_efficiency=Avg('efficiency_score'),
        total_customers=Count('customer_name', distinct=True)
    ).order_by('-total_loads')
    
    for row_num, item in enumerate(driver_data, 2):
        row_data = [
            item['driver_name'],
            item['truck_number'],
            item['total_loads'],
            round(item['total_distance'] or 0, 2),
            round(item['avg_efficiency'] or 0, 2),
            item['total_customers'],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)


def reports_view(request):
    """Display reports and analytics"""
    # Get summary statistics for reports
    total_loads = TruckPerformanceData.objects.count()
    
    # Monthly breakdown
    monthly_breakdown = TruckPerformanceData.objects.values('month_name').annotate(
        total_loads=Count('id'),
        avg_efficiency=Avg('efficiency_score'),
        total_distance=Sum('total_distance')
    ).order_by('create_date')
    
    # Top performers
    top_drivers = TruckPerformanceData.objects.values('driver_name').annotate(
        total_loads=Count('id'),
        avg_efficiency=Avg('efficiency_score')
    ).order_by('-avg_efficiency')[:10]
    
    top_trucks = TruckPerformanceData.objects.values('truck_number').annotate(
        total_loads=Count('id'),
        avg_efficiency=Avg('efficiency_score')
    ).order_by('-avg_efficiency')[:10]
    
    context = {
        'total_loads': total_loads,
        'monthly_breakdown': monthly_breakdown,
        'top_drivers': top_drivers,
        'top_trucks': top_trucks,
    }
    
    return render(request, 'dashboard/reports.html', context)
