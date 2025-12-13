from django.utils.timezone import is_aware, make_naive
import datetime
def to_naive(dt):
    if dt is None:
        return ''
    if isinstance(dt, datetime.datetime):
        if is_aware(dt):
            dt = make_naive(dt)
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(dt, datetime.date):
        return dt.strftime('%Y-%m-%d')
    return str(dt)
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import TruckPerformanceData
from typing import Any

def export_excel_report(request) -> Any:
    """
    Generate and return a combined Excel report of all processed truck performance data.
    """

    data = TruckPerformanceData.objects.all().order_by('create_date', 'load_number')
    if not data.exists():
        messages.error(request, 'No data available to export. Please upload and process CSV files first.')
        return redirect('dashboard:bulk_upload')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Truck Productivity Data'

    header_names = [
        'Create Date', 'Month Name', 'Transporter', 'Load Number', 'Mode Of Capture', 'Driver Name', 'Vehicle Reg', 'Customer Name',
        'Vol Hl', 'Invoice Number', 'Mwarehouse', 'Budgeted Kms', 'PlannedDistanceToCustomer', 'Actual Km', 'Km Deviation', 'Comment', 'Clockin Time',
        'Planned Departure Time', 'Dj Departure Time', 'Departure Deviation Min', 'Ave Departure', 'Comment Ave Departure',
        'Arrival At Customer', 'Departure Time From Customer', 'Service Time At Customer', 'Comment Tat', 'Arrival At Depot',
        'Clock Out', 'Ave Arrival Time', 'Comment Ave Arrival Time', 'Actual Days In Route', 'Bud Days In Route',
        'Days In Route Deviation', 'Total Hour Route', 'Driver Rest Hours In Route', 'Total Wh', 'Tlp',
        'D1', 'D2', 'D3', 'D4', 'Comment Ave Tir'
    ]
    from openpyxl.styles import PatternFill
    ws.append(header_names)
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Build robust mapping from (driver_name, load_number) to truck_number from depot_departures
    depot_departures = TruckPerformanceData.objects.filter(csv_upload__upload_type='depot_departures')
    mapping = {}
    for dep in depot_departures:
        driver = dep.driver_name.strip().lower() if dep.driver_name else ''
        load = dep.load_number.strip().lower() if dep.load_number else ''
        truck = dep.truck_number.strip() if dep.truck_number else ''
        if driver and load and truck and truck.lower() != 'unknown':
            mapping[(driver, load)] = truck
        # Also allow mapping by driver only if unique
        if driver and truck and truck.lower() != 'unknown':
            mapping[(driver, '')] = truck
        if load and truck and truck.lower() != 'unknown':
            mapping[('', load)] = truck

    for item in data:
        # Map Vehicle Reg using (driver_name, load_number)
        driver = item.driver_name.strip().lower() if item.driver_name else ''
        load = item.load_number.strip().lower() if item.load_number else ''
        # Try full match, then driver only, then load only, then fallback
        truck_number = (
            mapping.get((driver, load))
            or mapping.get((driver, ''))
            or mapping.get(('', load))
            or (item.truck_number if item.truck_number and str(item.truck_number).strip().lower() != 'unknown' else '')
        )

        row = [
            item.create_date.strftime('%Y-%m-%d') if item.create_date else '',
            item.month_name or '',
            item.transporter or '',
            item.load_number or '',
            item.mode_of_capture or '',
            item.driver_name or '',
            truck_number or item.truck_number or '',
            item.customer_name or '',
            item.tlp_vol_hl or '',  # Vol Hl mapped
            item.load_number or '',  # Invoice Number uses load_number
            'OM',  # Mwarehouse set to OM
            item.budgeted_kms or '',
            getattr(item, 'PlannedDistanceToCustomer', None) or '',
            '',  # Actual Km (not mapped in model)
            item.km_deviation or '',
            getattr(item, 'comment', '') or '',
            to_naive(item.clockin_time) if item.clockin_time else '',
            to_naive(item.planned_departure_time) if item.planned_departure_time else '',
            to_naive(item.dj_departure_time) if item.dj_departure_time else '',
            item.departure_deviation_min or '',
            item.ave_departure or '',
            '',
            to_naive(item.arrival_at_customer) if item.arrival_at_customer else '',
            to_naive(item.departure_time_from_customer) if item.departure_time_from_customer else '',
            item.service_time_at_customer or '',
            '',
            to_naive(item.arrival_at_depot) if item.arrival_at_depot else '',
            to_naive(item.clock_out) if item.clock_out else '',
            item.ave_arrival_time or '',
            '',
            item.actual_days_in_route or '',
            item.bud_days_in_route or '',
            item.days_in_route_deviation or '',
            item.total_hour_route or '',
            item.driver_rest_hours_in_route or '',
            item.total_wh or '',
            item.tlp_vol_hl or '',
            item.D1 or '',
            item.D2 or '',
            item.D3 or '',
            item.D4 or '',
            '',
        ]
        ws.append(row)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Truck_Productivity_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
